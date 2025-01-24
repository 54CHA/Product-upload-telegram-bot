import asyncio
import openpyxl
import io
import os
import aiohttp
from dotenv import load_dotenv
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, CallbackQueryHandler, ContextTypes, filters
import nest_asyncio

nest_asyncio.apply()

load_dotenv()

TELEGRAM_TOKEN = os.getenv('TELEGRAM_TOKEN')
STRAPI_API_TOKEN = os.getenv('STRAPI_API_TOKEN')
STRAPI_API_URL = os.getenv('STRAPI_API_URL')

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [
        [
            InlineKeyboardButton("📥 Загрузить товары", callback_data='upload_products'),
            InlineKeyboardButton("📄 Скачать шаблон", callback_data='download_template')
        ]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text(
        'Добро пожаловать! Выберите действие:',
        reply_markup=reply_markup
    )

async def button(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    if query.data == 'upload_products':
        await query.edit_message_text(
            text="Пожалуйста, отправьте Excel файл с товарами.\n\n"
                 "Excel файл должен содержать следующие столбцы:\n"
                 "A: Название\n"
                 "B: Артикул\n"
                 "C: Описание\n"
                 "D: ID категории\n"
                 "E: ID подкатегории\n"
                 "F: ID бренда\n"
                 "G: ID модели\n"
                 "H: ID модификации\n"
                 "I: Спецификации (краткие)\n"
                 "J: Спецификации (подробные)\n"
                 "K: Ссылка где купить"
        )
    elif query.data == 'download_template':
        await create_and_send_template(update.callback_query.message)

async def process_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        file = await context.bot.get_file(update.message.document.file_id)
        file_bytes = await file.download_as_bytearray()
        
        products = extract_data_from_excel(file_bytes)
        
        if not products:
            await update.message.reply_text("В Excel файле не найдено товаров или произошла ошибка при обработке.")
            return

        await update.message.reply_text(f"Найдено {len(products)} товаров. Начинаю загрузку в Strapi...")
        
        async with aiohttp.ClientSession() as session:
            success_count = 0
            duplicate_count = 0
            error_count = 0
            
            for product in products:
                result = await create_product_in_strapi(session, product, None)
                if result['success']:
                    success_count += 1
                    await update.message.reply_text(f"✅ Создан: {product['name']}")
                else:
                    if result['reason'] == 'duplicate':
                        duplicate_count += 1
                        await update.message.reply_text(f"⚠️ Пропущен дубликат: {product['name']}")
                    else:
                        error_count += 1
                        await update.message.reply_text(f"❌ Ошибка создания: {product['name']}")
        
        await update.message.reply_text(
            f"Загрузка завершена!\n"
            f"✅ Успешно создано: {success_count} товаров\n"
            f"⚠️ Пропущено дубликатов: {duplicate_count} товаров\n"
            f"❌ Ошибок: {error_count} товаров"
        )
    
    except Exception as e:
        await update.message.reply_text(f"Произошла ошибка: {str(e)}")

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        'Пожалуйста, отправьте Excel файл с данными о товарах.'
    )

async def create_product_in_strapi(session, product_data, image_id):
    try:
        headers = {
            'Authorization': f'Bearer {STRAPI_API_TOKEN}',
            'Content-Type': 'application/json'
        }

        # URL encode the article number for the query
        encoded_article = product_data["article"].replace(' ', '%20')
        
        # Check if product with this article number already exists
        async with session.get(
            f'{STRAPI_API_URL}/api/catalog-products?filters[articleNumber][$eq]={encoded_article}',
            headers=headers
        ) as response:
            if response.status == 200:
                existing_products = await response.json()
                if existing_products.get('data') and len(existing_products['data']) > 0:
                    return {'success': False, 'reason': 'duplicate'}
            else:
                print(f"Error checking for duplicates: {await response.text()}")

        slug = product_data['name'].lower().replace(' ', '-')
        
        # Convert specifications to the correct format
        specs = product_data.get('specifications', [])
        detailed_specs = product_data.get('detailedSpecifications', [])

        data = {
            "data": {
                "name": product_data['name'],
                "slug": slug,
                "articleNumber": product_data['article'],
                "description": product_data['description'],
                "specifications": specs,
                "detailedSpecifications": detailed_specs,
                "whereToBuyLink": product_data.get('whereToBuyLink', ""),
                "publishedAt": None
            }
        }

        # Add relations with proper format for Strapi v4 manyToOne relations
        if product_data.get('category'):
            data["data"]["category"] = {"id": product_data['category']}
        if product_data.get('subcategory'):
            data["data"]["subcategory"] = {"id": product_data['subcategory']}
        if product_data.get('brand'):
            data["data"]["brand"] = {"id": product_data['brand']}
        if product_data.get('model'):
            data["data"]["model"] = {"id": product_data['model']}
        if product_data.get('modification'):
            data["data"]["modification"] = {"id": product_data['modification']}

        # Add image if provided and not None
        if image_id:
            data["data"]["images"] = [image_id]

        print(f"Sending data to Strapi: {data}")

        async with session.post(
            f'{STRAPI_API_URL}/api/catalog-products',
            json=data,
            headers=headers
        ) as response:
            response_text = await response.text()
            print(f"Response from Strapi: {response_text}")
            if response.status not in [200, 201]:
                return {'success': False, 'reason': 'api_error'}
            return {'success': True}
    except Exception as e:
        print(f"Error creating product: {e}")
        return {'success': False, 'reason': 'exception', 'error': str(e)}

def extract_data_from_excel(excel_bytes):
    temp_file = 'temp_excel.xlsx'
    with open(temp_file, 'wb') as f:
        f.write(excel_bytes)

    try:
        workbook = openpyxl.load_workbook(temp_file)
        sheet = workbook.active

        products_data = []
        for row_idx in range(2, sheet.max_row + 1):
            if not sheet.cell(row=row_idx, column=1).value:
                continue

            # Process specifications (column I)
            spec_value = str(sheet.cell(row=row_idx, column=9).value or '')
            specs = []
            if spec_value:
                spec_parts = spec_value.split(',')
                for i, part in enumerate(spec_parts):
                    if ':' in part:
                        name, value = part.split(':', 1)
                        specs.append({
                            "label": name.strip(),
                            "value": value.strip()
                        })
                    else:
                        specs.append({
                            "label": f"Specification {i+1}",
                            "value": part.strip()
                        })

            # Process detailed specifications (column J)
            detailed_spec_value = str(sheet.cell(row=row_idx, column=10).value or '')
            detailed_specs = []
            if detailed_spec_value:
                detailed_spec_parts = detailed_spec_value.split(',')
                for i, part in enumerate(detailed_spec_parts):
                    if ':' in part:
                        name, value = part.split(':', 1)
                        detailed_specs.append({
                            "label": name.strip(),
                            "value": value.strip()
                        })
                    else:
                        detailed_specs.append({
                            "label": f"Specification {i+1}",
                            "value": part.strip()
                        })

            # Ensure at least one specification exists for both
            if not specs:
                specs = [{
                    "label": "General",
                    "value": "Not specified"
                }]
            if not detailed_specs:
                detailed_specs = [{
                    "label": "General",
                    "value": "Not specified"
                }]

            product = {
                'name': str(sheet.cell(row=row_idx, column=1).value or '').strip(),
                'article': str(sheet.cell(row=row_idx, column=2).value or '').strip(),
                'description': str(sheet.cell(row=row_idx, column=3).value or '').strip(),
                'category': int(sheet.cell(row=row_idx, column=4).value or 0),
                'subcategory': int(sheet.cell(row=row_idx, column=5).value or 0),
                'brand': int(sheet.cell(row=row_idx, column=6).value or 0),
                'model': int(sheet.cell(row=row_idx, column=7).value or 0),
                'modification': int(sheet.cell(row=row_idx, column=8).value or 0),
                'specifications': specs,
                'detailedSpecifications': detailed_specs,
                'whereToBuyLink': str(sheet.cell(row=row_idx, column=11).value or '').strip()
            }
            
            if product['name'] and product['article'] and product['category'] and product['whereToBuyLink']:
                products_data.append(product)
            else:
                print(f"Skipping row {row_idx} due to missing required fields")

        print(f"Successfully processed {len(products_data)} products from Excel")
        return products_data

    except Exception as e:
        print(f"Error processing Excel file: {e}")
        return []
    finally:
        if os.path.exists(temp_file):
            os.remove(temp_file)

async def create_and_send_template(message):
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        
        headers = ["Название", "Артикул", "Описание", "ID категории", "ID подкатегории", 
                  "ID бренда", "ID модели", "ID модификации", "Спецификации (краткие)", 
                  "Спецификации (подробные)", "Ссылка где купить"]
        
        # Add headers
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)
        
        # Add example row
        example_data = [
            "Тормозной диск передний",  # Название
            "BD-12345",                 # Артикул
            "Высококачественный тормозной диск для передней оси",  # Описание
            "1",                        # ID категории
            "2",                        # ID подкатегории
            "3",                        # ID бренда
            "4",                        # ID модели
            "5",                        # ID модификации
            "Диаметр:280мм, Толщина:22мм",  # Спецификации (краткие)
            "Диаметр:280мм, Толщина:22мм, Тип:Вентилируемый, Покрытие:С покрытием",  # Спецификации (подробные)
            "https://example.com/product"  # Ссылка где купить
        ]
        
        for col, value in enumerate(example_data, 1):
            sheet.cell(row=2, column=col, value=value)
        
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
        
        # Save template
        temp_file = 'template.xlsx'
        workbook.save(temp_file)
        
        # Send template
        await message.reply_document(
            document=open(temp_file, 'rb'),
            filename='template.xlsx',
            caption="✅ Шаблон для загрузки товаров"
        )
        
        os.remove(temp_file)
        
        await message.reply_text(
            "Используйте этот шаблон для подготовки данных.\n"
            "После заполнения отправьте файл боту для загрузки товаров."
        )
        
    except Exception as e:
        print(f"Error creating template: {e}")
        await message.reply_text(
            "❌ Произошла ошибка при создании шаблона.\n"
            "Попробуйте еще раз используя /start"
        )

def main():
    application = Application.builder().token(TELEGRAM_TOKEN).build()

    application.add_handler(CommandHandler("start", start))
    application.add_handler(CallbackQueryHandler(button))
    application.add_handler(MessageHandler(filters.Document.ALL, process_excel))
    application.add_handler(MessageHandler(filters.TEXT, handle_message))

    print("Starting bot...")
    application.run_polling()

if __name__ == '__main__':
    main()
